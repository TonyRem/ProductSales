import os
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

from data.constants import DATA_WHS
from scripts.format_files import apply_formatting

def custom_function(article, whs):
    if 'MAC75' in str(whs) or 'MAK75' in str(whs):
        if str(article).startswith('7'):
            return str(article)
        return str(article)[:5]
    else:
        if str(article).startswith('Z'):
            return str(article)
        return str(article)[:5]

def format_sales(sales_path, stock_path, merge_cells, result_file_path):
    sales_path = Path(sales_path)
    stock_path = Path(stock_path)
   
    sales_wb = load_workbook(sales_path)
    stock_wb = load_workbook(stock_path)
    sales_sheet = sales_wb.active
    stock_sheet = stock_wb.active
    
    sales_data = []
    for row in sales_sheet.iter_rows(values_only=True):
        sales_data.append(row)

    stock_data = []
    for row in stock_sheet.iter_rows(values_only=True):
        stock_data.append(row)

    sales_df = pd.DataFrame(sales_data[1:], columns=sales_data[0])
    stock_df = pd.DataFrame(stock_data[1:], columns=stock_data[0])
    whs_df = pd.DataFrame(DATA_WHS, columns=['WHS', 'WHS code'])

    stock_df['Free for sales'] = stock_df['Free for sales'].fillna(0)
    stock_df['Free for sales'] = stock_df['Free for sales'].astype(int)
    stock_df.rename(columns={'SKU': 'Stock_SKU'}, inplace=True)
   
    grouped_stock = stock_df.groupby(
        ['WHS code', 'Product article', 'Stock_SKU', ' EC Description', 'Product group']
    )['Free for sales'].sum().reset_index()

    merged_df = pd.merge(sales_df, whs_df, on='WHS', how='left')
    merged_df = pd.merge(
        merged_df,
        grouped_stock,
        how='outer',
        left_on=['WHS code', 'Article', 'SKU'],
        right_on=['WHS code', 'Product article', 'Stock_SKU']
    )

    merged_df['Product article'] = merged_df['Product article'].fillna(
        merged_df['Article']
    )
    merged_df['Article'] = merged_df['Article'].fillna(
        merged_df['Product article']
    )
    merged_df['Product description'] = merged_df['Product description'].fillna(
        merged_df[' EC Description']
    )
    merged_df['SKU'] = merged_df['SKU'].fillna(
        merged_df['Stock_SKU']
    )
    merged_df['Product type'] = merged_df['Product type'].fillna(
        merged_df['Product group']
    )
    merged_df['Date'].fillna(method='ffill', inplace=True)

    whs_mapping = whs_df.set_index('WHS code')['WHS'].to_dict()

    # Заполнение отсутствующих значений в столбце WHS на основе соответствующих значений из словаря
    merged_df['WHS'] = merged_df['WHS'].fillna(merged_df['WHS code'].map(whs_mapping))
    merged_df['WHS'] = merged_df['WHS'].fillna(merged_df['WHS code'])
    merged_df['Date'] = pd.to_datetime(
        merged_df['Date'],
        format='%d/%m/%Y %H:%M'
    )
    merged_df['Week'] = merged_df['Date'].dt.strftime('%Y-%U')
    merged_df['Units'] = merged_df['Units'].fillna(0)
    merged_df['Units'] = merged_df['Units'].astype(int)
    merged_df['Free for sales'] = merged_df['Free for sales'].fillna(0)

    grouped_data = merged_df.groupby(
        ['Product type','Article', 'SKU', 'Week', 'Product description',
         'WHS', 'WHS code', 'Free for sales']
    )['Units'].sum().reset_index()

    pivot_table = grouped_data.pivot_table(
        index=['Product type', 'Article', 'SKU', 'Product description',
               'WHS', 'WHS code', 'Free for sales'],
        columns=['Week'],
        values='Units',
        fill_value=0
    )
    pivot_table.reset_index(inplace=True)
    pivot_table.columns.name = None

    average_sales_by_week = pivot_table.copy()

    sales_columns = pivot_table.columns[7:]
    average_sales_by_week['AVG Sales'] = pivot_table[sales_columns].mean(axis=1).round(2)
    average_sales_by_week['Days of Stock'] = (average_sales_by_week['Free for sales'] / average_sales_by_week['AVG Sales'] * 7).round()
    average_sales_by_week.insert(1, 'SPN', average_sales_by_week.apply(lambda row: custom_function(row['Article'], row['WHS']), axis=1))

    organized_excel_file = result_file_path
    
    with pd.ExcelWriter(organized_excel_file, engine='openpyxl') as writer:
        average_sales_by_week.to_excel(
            writer,
            sheet_name='Sheet1',
            merge_cells=merge_cells)

    formatted_workbook = load_workbook(result_file_path)
    formatted_worksheet = formatted_workbook.active
    apply_formatting(formatted_worksheet)
    formatted_workbook.save(result_file_path)

    os.remove(sales_path)
    os.remove(stock_path)
