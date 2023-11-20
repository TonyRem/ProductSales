import os
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

def merge_matrix():
    #sbw_path = Path(sbw_path)
    sbw_path = Path('/Users/anton/Desktop/matrix/temp/SalesByWeek_17.08_10.38.xlsx')
    matrix_path = Path('/Users/anton/Desktop/matrix/data/matrix_articles.xlsx')
   
    sbw_wb = load_workbook(sbw_path)
    matrix_wb = load_workbook(matrix_path)
    sales_sheet = sbw_wb.active
    stock_sheet = matrix_wb.active
    
    sbw_data = []
    for row in sales_sheet.iter_rows(values_only=True):
        sbw_data.append(row)

    matrix_data = []
    for row in stock_sheet.iter_rows(values_only=True):
        matrix_data.append(row)

    sbw_df = pd.DataFrame(sbw_data[1:], columns=sbw_data[0])
    matrix_df = pd.DataFrame(matrix_data[1:], columns=matrix_data[0])
    
    # Объединение данных
    merged_df = matrix_df.merge(sbw_df, how="left", left_on="Article", right_on="Article")
    
    # Заполнение пропусков нулями
    #merged_df.fillna(0, inplace=True)
    organized_excel_file = 'hoba.xlsx'
    with pd.ExcelWriter(organized_excel_file, engine='openpyxl') as writer:
        merged_df.to_excel(
            writer,
            sheet_name='Sheet1')

merge_matrix()