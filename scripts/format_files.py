from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def apply_formatting(worksheet):
    column_widths = {
        'A': 5, 'B': 15, 'C': 10, 'D': 15, 'E': 15, 'F': 50, 'G': 15, 'H': 13
    }
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2E75B6",
        end_color="2E75B6",
        fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color="000000"),
                                     right=Side(style='thin', color="000000"),
                                     top=Side(style='thin', color="000000"),
                                     bottom=Side(style='thin', color="000000"))
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill(
                    start_color="F2F2F2",
                    end_color="F2F2F2",
                    fill_type="solid"
                )
                
